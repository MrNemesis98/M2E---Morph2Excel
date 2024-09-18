"""
Copyright © MrNemesis98, GitHub, 2024

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the
Software. The software is provided “as is”, without warranty of any kind, express or implied, including but not
limited to the warranties of merchantability, fitness for a particular purpose and noninfringement.
In no event shall the author(s) or copyright holder(s) be liable for any claim, damages or other liability, whether
in an action of contract, tort or otherwise, arising from, out of or in connection with the software or the use or
other dealings in the software.
"""

import os
import sys
import time

from openpyxl import load_workbook
import win32gui
import win32con
import win32api

import console_text_management as CTM
import console_assistance as CA
import savedata_manager as SDM
import notification_sound_player as NSP

# system variables
workbook = None
worksheet = None
workbook_title = None
hdlp_start = False
hdlp_doc = False
comparison_counter = 1
tip_need_counter = 0
excel_row = 0
entries_list = None
database_is_installed = False

# data to change for every release
m2e_version = "2024.1"
supporter = "Till Preidt ~ GitHub/MrNemesis98"
support_email = "s2tiprei@uni-trier.de / till.p2.tp@gmail.com"


def set_system_variables_to_default():
    global database_version_date
    global database_version_description
    global term_output_diplomacy
    global oneline_output_format
    global headline_printing
    global alphabetical_output, abc_output_ascending
    global auto_scan_filters
    global output_detail_level
    global system_sound_level

    SDM.set_database_version_date("")
    SDM.set_database_version_description("")
    SDM.set_term_output_diplomacy("3")
    SDM.set_one_line_output(True)
    SDM.set_headline_printing("2")
    SDM.set_alphabetical_output(True, True)
    SDM.set_auto_scan_filters("Noun, Verb, Adjective, Adverb, Preposition, Phrase")
    SDM.set_output_detail_level("3")
    SDM.set_system_sound_level("3")

    database_version_date = SDM.get_database_version_date()
    database_version_description = SDM.get_database_version_description()
    term_output_diplomacy = SDM.get_term_output_diplomacy()
    oneline_output_format = SDM.get_one_line_output()
    headline_printing = SDM.get_headline_printing()
    alphabetical_output, abc_output_ascending = SDM.get_alphabetical_output()
    auto_scan_filters = SDM.get_auto_scan_filters()
    output_detail_level = SDM.get_output_detail_level()
    system_sound_level = SDM.get_system_sound_level()


def get_screen_size():
    screen_width = win32api.GetSystemMetrics(win32con.SM_CXSCREEN)
    screen_height = win32api.GetSystemMetrics(win32con.SM_CYSCREEN)
    return screen_width, screen_height


def set_console_fullscreen():
    hwnd = win32gui.GetForegroundWindow()

    screen_width, screen_height = get_screen_size()

    win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, win32con.WS_POPUP)
    win32gui.SetWindowPos(hwnd, win32con.HWND_TOP, 0, 0, screen_width, screen_height,
                          win32con.SWP_FRAMECHANGED | win32con.SWP_SHOWWINDOW)


def disable_resize_and_buttons():
    hwnd = win32gui.GetForegroundWindow()

    style = win32gui.GetWindowLong(hwnd, win32con.GWL_STYLE)

    style &= ~win32con.WS_CAPTION
    style &= ~win32con.WS_THICKFRAME
    style &= ~win32con.WS_MAXIMIZEBOX
    style &= ~win32con.WS_SYSMENU

    win32gui.SetWindowLong(hwnd, win32con.GWL_STYLE, style)
    win32gui.SetWindowPos(hwnd, None, 0, 0, 0, 0,
                          win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_NOZORDER | win32con.SWP_FRAMECHANGED)


def check_database_installation():
    global database_version_date
    global database_version_description
    global system_sound_level
    global database_is_installed
    time.sleep(2)

    CTM.draw("\r\tChecking database status...")
    time.sleep(1.5)

    if not CA.database_installation_confirmed(right_after_program_start=True):

        SDM.set_database_version_date("")
        SDM.set_database_version_description("")

        CTM.unblock_input()
        i = input("\n\n\t\tanswer: ")
        CTM.block_input()

        if i == "start!" or i == "start" or i == "start1":

            NSP.play_accept_sound() if system_sound_level == 3 else None
            url = "https://zenodo.org/record/5172857/files/wiki_morph.json?download=1"
            normal = CA.download_database(url=url, directly_after_start=True)

            if normal:
                current_size = os.path.getsize("src/database/wiki_morph.json")
                current_size = int(current_size / (1024 * 1024))
                SDM.set_current_size(current_size)
                database_version_description = SDM.get_database_version_description()
                database_version_date = SDM.get_database_version_date()
                print("\r\t\033[92mDownload completed sucessfully!\033[0m (" + str(current_size) +
                      " MB)\n\tThe latest version of wikimorph is now installed on your device."
                      "\n\n\t\33[94mTip:\33[0m In the \33[33msettings menu\33[0m you can manage this version and add a "
                      "description."
                      "\n\n\tPress \33[92menter\33[0m to continue.", end='', flush=True)
                NSP.play_start_sound() if system_sound_level >= 2 else None
                database_is_installed = True
                CTM.unblock_input()
                input()
                CTM.block_input()
            else:
                NSP.play_deny_sound() if system_sound_level == 3 else None
                sys.exit()

    else:
        database_is_installed = True


# preparing worksheet with first (general) headline
def prepare_worksheet():
    global workbook
    global worksheet
    global workbook_title
    global excel_row
    global hdlp_start
    global formatted_date

    workbook_title = CA.create_excel(fd=formatted_date)
    excel_row = 1
    workbook = load_workbook(workbook_title)
    worksheet = workbook.active
    worksheet, excel_row = CA.print_headlines(worksheet, excel_row, output_detail_level)
    hdlp_start = True

    return workbook, worksheet, excel_row, hdlp_start


def search_for_terms(log_title):

    # Excel file variables
    global workbook
    global worksheet
    global excel_row
    global entries_list
    global comparison_counter
    global hdlp_start
    global hdlp_doc

    # System variables
    global database_is_installed
    global database_version_date
    global database_version_description
    global term_output_diplomacy
    global oneline_output_format
    global headline_printing
    global alphabetical_output
    global abc_output_ascending
    global auto_scan_filters
    global output_detail_level
    global system_sound_level
    global tip_need_counter

    worksheet_generated = False
    open_excel_automatically = False
    comparison_workbooks = []
    print_main_menu_again = True

    if database_is_installed:
        entries_list = CA.load_database(version=m2e_version)
    CTM.clear_screen_backwards(down_to_row=1, delay=0)
    CA.print_opening(version=m2e_version, colour=False)
    time.sleep(1)

    stop = False

    while not stop:

        if print_main_menu_again:
            tip_need_counter = 0
            CTM.block_input()
            CTM.clear_screen_backwards(down_to_row=5)
            CA.print_main_menu()
            print_main_menu_again = False
            CTM.unblock_input()
        else:
            tip_need_counter += 1
            if tip_need_counter == 3:
                CA.print_manual_search_headline(tip=True)
                tip_need_counter = 0
            else:
                CA.print_manual_search_headline(tip=False)
        CTM.unblock_input()
        i = input("\n\t\33[97mSearch term: \33[92m").lower()
        CTM.block_input()
        print("\33[0m")
        CTM.clear_screen_backwards(down_to_row=5)

        if i == "exit!":
            CTM.draw("\033[91m" + "\n\tProgram terminated!\033[0m", clear=False)
            NSP.play_accept_sound() if system_sound_level == 3 else None
            time.sleep(1)
            stop = True
            if open_excel_automatically:
                CTM.draw("\033[33m" + "\n\tOpening search results...\033[0m", clear=False)
                time.sleep(2)
                os.system(f'start "" {workbook_title}')
            if comparison_counter != 1:
                CTM.draw("\033[94m" + "\n\tOpening comparison results...\033[0m", clear=False)
                time.sleep(2)
                for file in comparison_workbooks:
                    os.system(f'start "" {file}')
                    time.sleep(2)
            time.sleep(2)
            os.system('cls')
        elif i == "i!" or i == "?":
            CA.show_instructions()
            time.sleep(3)
            print_main_menu_again = True
        elif i == "v!":
            CA.show_version_description(version=m2e_version)
            CTM.unblock_input()
            i = input()
            CTM.block_input()
            print_main_menu_again = True
        elif i == "":
            print_main_menu_again = True

        # AUTOMATIC SCAN MODE ------------------------------------------------------------------------------------------
        elif i == "s!":

            if not worksheet_generated:
                workbook, worksheet, excel_row, hdlp_start = prepare_worksheet()
                worksheet_generated = True

            if CA.database_installation_confirmed():

                open_excel_automatically = True
                print("\n\t\033[38;5;130m- Automatic Scan Mode -\033[0m"
                      "\n\t\033[38;5;130m------------------------------------------------------------------------\033[0m")
                NSP.play_accept_sound() if system_sound_level == 3 else None

                excel_file_selected = False
                breakoff = False
                while not excel_file_selected:
                    time.sleep(1)
                    CTM.draw("\n\t\33[33mPlease select an excel file to scan for possible terms.\33[0m", clear=False)
                    time.sleep(1.5)
                    try:
                        file = CA.select_excel_file()
                        NSP.play_accept_sound() if system_sound_level == 3 else None

                        start_time = time.time()
                        terms, invalid_cases = CA.autoscan(file, duplicates=False,
                                                           abc=alphabetical_output, abc_ascending=abc_output_ascending)
                        end_time = time.time()
                        number_of_terms = len(terms) + len(invalid_cases)
                        number_of_valid_cases = len(terms)
                        status = ["\n\tExcel file: " + file,
                                  "\tFound terms: " + str(number_of_terms),
                                  "\tValid cases: " + str(number_of_valid_cases),
                                  "\tPos filters: " + str(auto_scan_filters)]

                        time.sleep(1)
                        CTM.clear_screen_backwards(down_to_row=8)
                        CTM.stack(status, clear=False)
                        CTM.draw("\n\t", CA.measure_time(start_time, end_time, search=False))
                        time.sleep(3)
                        if not invalid_cases == [] and len(invalid_cases) <= 10:
                            print("\n\t\033[91mWarning:\033[0m Scanned file contains terms that are invalid inputs!"
                                  "\n\n\tFor searching within the wikimorph database "
                                  "the following terms will be ignored:"
                                  "\n\t\t", str(invalid_cases))
                            NSP.play_deny_sound() if system_sound_level >= 2 else None
                        elif not invalid_cases == [] and len(invalid_cases) > 10:
                            print("\n\t\033[91mWarning:\033[0m Scanned file contains terms that are invalid inputs!"
                                  "\n\n\tThese terms will be ignored for searching within the wikimorph database."
                                  "\n\tThe amount of invalid terms is too large to be displayed here.")
                            NSP.play_deny_sound() if system_sound_level >= 2 else None
                        CTM.unblock_input()
                        i = input("\n\tPress \033[92menter\033[0m or type in anything to "
                                  "\033[92mstart\033[0m the search. "
                                  '\n\tType in \033[33mfile!\033[0m to select a different file.'
                                  '\n\tType in \033[91mexit!\033[0m to cancel the automatic search.'
                                  '\n\n\ta'
                                  'nswer: ')
                        CTM.block_input()
                        if i == "exit!":
                            excel_file_selected = True
                            breakoff = True
                        elif i == "file!":
                            NSP.play_accept_sound() if system_sound_level == 3 else None
                        else:
                            excel_file_selected = True

                    except Exception:
                        CTM.clear_screen_backwards(down_to_row=8)
                        print("\n\t\033[91mWarning:\033[0m No file selected!")
                        NSP.play_deny_sound() if system_sound_level >= 2 else None

                if not breakoff:
                    CTM.draw("\n\t\033[92mThe valid terms will now be searched in the database.\033[0m", clear=False)
                    NSP.play_accept_sound() if system_sound_level == 3 else None
                    time.sleep(2)

                    # for preventing double headline printing either at the beginning of the Excel
                    # (conflict with hdlp_start) or before every term (in case hdlp=3)
                    # only executes the "for every new doc" headline printing (hdlp=2)
                    # the headline printing for every term (hdlp=3) is defined below in the basic search section and
                    # in the search_and_output function of console_assistance.py
                    if headline_printing == 2:
                        if not hdlp_start:
                            worksheet, excel_row = CA.print_headlines(worksheet, excel_row, output_detail_level)
                            hdlp_doc = True
                        else:
                            hdlp_start = False

                    start_time = time.time()
                    log = open(log_title, "a", encoding="utf-8")
                    if term_output_diplomacy == 1:
                        for x in range(number_of_valid_cases):
                            term = terms[x]
                            CTM.clear_screen_backwards(down_to_row=13)
                            progress = format(100 * (x / number_of_valid_cases), ".2f")
                            print("\n\t\033[38;5;130mSearching for terms...\033[0m"
                                  "\n\tCurrent term: " + term +
                                  "\t\tProgress: \33[38;5;130m" + str(progress) + "%\33[0m")

                            worksheet, excel_row, log_output, \
                                hdlp_start, hdlp_doc = CA.search_and_output(worksheet=worksheet,
                                                                            excel_row=excel_row,
                                                                            pos_filters=auto_scan_filters,
                                                                            term=term,
                                                                            entries_list=entries_list,
                                                                            only_found_terms=True,
                                                                            only_not_found_terms=False,
                                                                            multiline_output=not oneline_output_format,
                                                                            output_detail_level=output_detail_level,
                                                                            headline_printing=headline_printing,
                                                                            hdlp_start=hdlp_start,
                                                                            hdlp_doc=hdlp_doc)
                            log.write("\n\n" + log_output)
                            time.sleep(.1)

                    elif term_output_diplomacy == 2:
                        for x in range(number_of_valid_cases):
                            term = terms[x]
                            CTM.clear_screen_backwards(down_to_row=13)
                            progress = format(100 * (x / number_of_valid_cases), ".2f")
                            print("\n\t\033[38;5;130mSearching for terms...\033[0m"
                                  "\n\tCurrent term: " + term + "\t\tProgress: \33[38;5;130m" + str(progress) + "%\33[0m")

                            worksheet, excel_row, log_output, \
                                hdlp_start, hdlp_doc = CA.search_and_output(worksheet=worksheet,
                                                                            excel_row=excel_row,
                                                                            pos_filters=auto_scan_filters,
                                                                            term=term,
                                                                            entries_list=entries_list,
                                                                            only_found_terms=False,
                                                                            only_not_found_terms=True,
                                                                            multiline_output=not oneline_output_format,
                                                                            output_detail_level=output_detail_level,
                                                                            headline_printing=headline_printing,
                                                                            hdlp_start=hdlp_start,
                                                                            hdlp_doc=hdlp_doc)
                            log.write("\n\n" + log_output)
                            time.sleep(.1)

                    else:
                        for x in range(number_of_valid_cases):
                            term = terms[x]
                            CTM.clear_screen_backwards(down_to_row=13)
                            progress = format(100 * (x / number_of_valid_cases), ".2f")
                            print("\n\t\033[38;5;130mSearching for terms...\033[0m"
                                  "\n\tCurrent term: " + term + "\t\tProgress: \33[38;5;130m" + str(progress) + "%\33[0m")

                            worksheet, excel_row, log_output, \
                                hdlp_start, hdlp_doc = CA.search_and_output(worksheet=worksheet,
                                                                            excel_row=excel_row,
                                                                            pos_filters=auto_scan_filters,
                                                                            term=term,
                                                                            entries_list=entries_list,
                                                                            only_found_terms=False,
                                                                            only_not_found_terms=False,
                                                                            multiline_output=not oneline_output_format,
                                                                            output_detail_level=output_detail_level,
                                                                            headline_printing=headline_printing,
                                                                            hdlp_start=hdlp_start,
                                                                            hdlp_doc=hdlp_doc)
                            log.write("\n\n" + log_output)
                            time.sleep(.1)

                    log.close()
                    workbook.save(workbook_title)
                    end_time = time.time()
                    CTM.clear_screen_backwards(down_to_row=13)
                    print("\n\t\033[92mProcess finished!\033[0m"
                          "\n\n\t", CA.measure_time(start_time, end_time))
                    NSP.play_accept_sound() if system_sound_level >= 2 else None
                    time.sleep(3)
                    CTM.clear_screen_backwards(down_to_row=14)
                    CTM.draw("\n\n\t" + CA.measure_time(start_time, end_time), clear=True)
                    CTM.draw("\n\n\t\33[92mResults were saved.\33[0m", clear=True)
                    time.sleep(1)
                    CTM.draw("\n\tReturning to main menu...", clear=True)
                    time.sleep(1)

                else:
                    CTM.clear_screen_backwards(down_to_row=8)
                    print("\n\t\033[91mProcess cancelled!\033[0m")
                    NSP.play_deny_sound() if system_sound_level == 3 else None
                    time.sleep(1)
                    CTM.draw("\n\tReturning to main menu...")
                print_main_menu_again = True
                time.sleep(3)

            else:
                SDM.set_database_version_date("")
                SDM.set_database_version_description("")
                print_main_menu_again = True
                CTM.unblock_input()
                input()
                CTM.block_input()

        # COMPARISON MODE ----------------------------------------------------------------------------------------------
        elif i == "c!":
            CTM.clear_screen_backwards(down_to_row=5)
            status = "\n\t\033[94m- Comparison Mode -\033[0m" \
                     "\n\t\033[94m------------------------------------------------------------------------\033[0m"
            print(status)
            NSP.play_accept_sound() if system_sound_level == 3 else None

            excel_file_1_selected = False
            breakoff = False
            while not excel_file_1_selected:
                CTM.draw("\n\tPlease select two excel files you want to compare.", clear=False)
                time.sleep(2)
                CTM.draw("\t\33[33mSelect excel file 1 now.\33[0m", clear=False)
                time.sleep(1)

                try:
                    file_1 = CA.select_excel_file()
                    NSP.play_accept_sound() if system_sound_level == 3 else None

                    start_time = time.time()
                    terms_1, invalid_terms_1 = CA.autoscan(file_1, test_for_invalides=False)
                    end_time = time.time()

                    number_of_terms_1 = len(terms_1)
                    CTM.stack(["\n\n\tSelected as file 1: " + file_1,
                               "\n\tFound terms: " + str(number_of_terms_1),
                               "\n" + CA.measure_time(start_time, end_time, search=False)])

                    time.sleep(2)

                    CTM.unblock_input()
                    i = input("\n\n\tPress \033[92menter\033[0m or type in anything to \033[92mselect file 2\033[0m. "
                              '\n\tType in \033[33mfile!\033[0m to select a different file.'
                              '\n\tType in \033[91mexit!\033[0m to cancel the comparison.'
                              '\n\n\tanswer: ')
                    CTM.block_input()
                    if i == "exit!":
                        excel_file_1_selected = True
                        breakoff = True
                    elif i == "file!":
                        CTM.clear_screen_backwards(down_to_row=8)
                        NSP.play_accept_sound() if system_sound_level == 3 else None
                    else:
                        excel_file_1_selected = True

                except Exception:
                    CTM.clear_screen_backwards(down_to_row=8)
                    print("\n\t\033[91mWarning:\033[0m No file as file 1 selected!")
                    NSP.play_deny_sound() if system_sound_level >= 2 else None
                    time.sleep(2)

            if not breakoff:
                excel_file_2_selected = False

                while not excel_file_2_selected:

                    CTM.clear_screen_backwards(down_to_row=8)
                    CTM.draw("\n\t\33[33mSelect excel file 2 now.\33[0m", clear=False)
                    time.sleep(1.5)

                    try:
                        file_2 = CA.select_excel_file()
                        NSP.play_accept_sound() if system_sound_level == 3 else None

                        start_time = time.time()
                        terms_2, invalid_terms_2 = CA.autoscan(file_2, test_for_invalides=False)
                        end_time = time.time()

                        number_of_terms_2 = len(terms_2)
                        CTM.stack(["\n\tSelected file 2: " + file_2,
                                   "\n\tFound terms: " + str(number_of_terms_2),
                                   "\n" + CA.measure_time(start_time, end_time, search=False)])

                        time.sleep(2)

                        CTM.unblock_input()
                        i = input("\n\n\tPress \033[92menter\033[0m or type in anything to \033[92mstart "
                                  "the comparison\033[0m. "
                                  '\n\tType in \033[33mfile!\033[0m to select a different file.'
                                  '\n\tType in \033[91mexit!\033[0m to cancel the comparison.'
                                  '\n\n\tanswer: ')
                        CTM.block_input()
                        if i == "exit!":
                            excel_file_2_selected = True
                            breakoff = True
                        elif i == "file!":
                            CTM.clear_screen_backwards(down_to_row=8)
                            NSP.play_accept_sound() if system_sound_level == 3 else None
                        else:
                            excel_file_2_selected = True

                    except Exception:
                        CTM.clear_screen_backwards(down_to_row=8)
                        print("\n\t\033[91mWarning:\033[0m No file as file 2 selected!")
                        NSP.play_deny_sound() if system_sound_level >= 2 else None
                        time.sleep(2)

            if not breakoff:
                CTM.clear_screen_backwards(down_to_row=8)
                CTM.draw("\n\tPreparing comparison...", clear=False)
                unique_terms_1 = []
                unique_terms_2 = []
                common_terms = []
                time.sleep(2)

                start_time = time.time()
                NSP.play_deny_sound() if system_sound_level >= 2 else None
                for x in range(number_of_terms_1):
                    term = terms_1[x]
                    progress = format(100 * ((x + 1) / number_of_terms_1), ".2f")
                    CTM.clear_screen_backwards(down_to_row=8)
                    print("\n\t\33[94mComparing terms from file 1...\33[0m"
                          "\n\tCurrent term: " + str(term) + "\t\tProgress: \33[94m" + str(progress) + "%\33[0m")
                    if term in terms_2:
                        common_terms.append(term)
                    else:
                        unique_terms_1.append(term)
                    time.sleep(.1)
                end_time = time.time()
                time.sleep(2)

                CTM.clear_screen_backwards(down_to_row=8)
                CTM.draw("\n\t\033[92mComparison of terms from file 1 finished!\033[0m", clear=False)
                CTM.draw("\n" + CA.measure_time(start_time, end_time, search=False, comparison=True), clear=False)
                NSP.play_deny_sound() if system_sound_level >= 2 else None
                CTM.unblock_input()
                input("\n\tType in anything to compare terms of file 2: ")
                CTM.block_input()
                time.sleep(1)

                start_time = time.time()
                NSP.play_accept_sound() if system_sound_level == 3 else None
                for x in range(number_of_terms_2):
                    term = terms_2[x]
                    progress = format(100 * ((x + 1) / number_of_terms_2), ".2f")
                    CTM.clear_screen_backwards(down_to_row=8)
                    print("\n\t\33[94mComparing terms from file 2...\33[0m"
                          "\n\tCurrent term: " + str(term) + "\t\tProgress: \33[94m" + str(progress) + "%\33[0m")
                    if term in terms_1:
                        common_terms.append(term)
                    else:
                        unique_terms_2.append(term)
                    time.sleep(.1)
                end_time = time.time()
                time.sleep(2)

                # eliminating duplicates in common terms list
                common_terms = list(set(common_terms))

                CTM.clear_screen_backwards(down_to_row=8)
                CTM.draw("\n\t\033[92mComparison of terms from file 2 finished!\033[0m", clear=False)
                CTM.draw("\n" + CA.measure_time(start_time, end_time, search=False, comparison=True), clear=False)
                NSP.play_deny_sound() if system_sound_level >= 2 else None
                CTM.unblock_input()
                input("\n\tType in anything to finish: ")
                CTM.block_input()

                CTM.clear_screen_backwards(down_to_row=8)
                CTM.draw("\n\t\033[92mComparing process finished!\033[0m", clear=False)
                NSP.play_accept_sound() if system_sound_level == 3 else None
                CTM.draw("\n\t\033[92mThe results were saved as an additional comparison excel file!\033[0m", clear=False)
                time.sleep(5)

                # generating new Excel file for comparison results exclusively
                results_wb_name = CA.create_comparison_result_excel(fd=formatted_date, counter=comparison_counter)
                results_workbook = load_workbook(results_wb_name)
                results_worksheet = results_workbook.active

                # saving results
                results_worksheet = CA.write_comparison_result_excel(worksheet=results_worksheet,
                                                                     file_1=file_1, file_2=file_2,
                                                                     list_of_terms_1=unique_terms_1,
                                                                     list_of_terms_2=unique_terms_2,
                                                                     common_terms_list=common_terms)
                results_workbook.save(results_wb_name)
                comparison_workbooks.append(results_wb_name)

                log = open(log_title, "a", encoding="utf-8")
                log_output = ("\t------------------------------------------------------------\n\n\t"
                              "Comparison mode accessed"
                              "\n\tFile 1: " + file_1 + "\n\tFile 2: " + file_2 + "\n")
                log.write("\n\n" + log_output)
                log.close()
                comparison_counter += 1

                CTM.clear_screen_backwards(down_to_row=8)
                CTM.draw("\n\tReturning to main menu...")
            else:
                CTM.clear_screen_backwards(down_to_row=8)
                CTM.draw("\n\t\033[91mProcess cancelled!\033[0m", clear=False)
                NSP.play_deny_sound() if system_sound_level == 3 else None
                CTM.draw("\n\tReturning to main menu...")

            time.sleep(4)
            print_main_menu_again = True

        # SETTINGS MODE ------------------------------------------------------------------------------------------------
        elif i in ["set!", "set1!", "set2!", "set3!", "set4!", "set5!",
                   "set6!", "set7!", "set8!"]:

            CTM.clear_screen_backwards(down_to_row=5, delay=0)
            print("\033[33m\n\t~ Settings Menu ~"
                  "\n\t-------------------------------------------------------------------------------\033[0m")
            NSP.play_accept_sound() if system_sound_level == 3 else None

            if i == "set!" or i == "set1!":
                setting_ctrl = 1
            elif i == "set2!":
                setting_ctrl = 2
            elif i == "set3!":
                setting_ctrl = 3
            elif i == "set4!":
                setting_ctrl = 4
            elif i == "set5!":
                setting_ctrl = 5
            elif i == "set6!":
                setting_ctrl = 6
            elif i == "set7!":
                setting_ctrl = 7
            else:
                setting_ctrl = 8

            # setting 1 (database version control center)

            while setting_ctrl <= 8:

                while setting_ctrl == 1:

                    # setting 1 (database version control center) ------------------------------------------------------

                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    database_version_date = SDM.get_database_version_date()
                    database_version_description = SDM.get_database_version_description()

                    CA.display_settings(setting=1, current_var=database_version_date,
                                        current_var_2=database_version_description)

                    CTM.unblock_input()
                    i = input("\n\n\tanswer: ")
                    CTM.block_input()

                    if i == "1":
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(setting=1, current_var="u1")
                        NSP.play_deny_sound() if system_sound_level >= 2 else None
                        CTM.draw("\n\t1. Press \33[92menter\33[0m to \33[92mstart the download\33[0m.",
                                 clear=False)
                        CTM.draw("\t2. Type in \33[91mexit!\33[0m to \33[91mreturn to settings menu\33[0m.",
                                 clear=False)

                        CTM.unblock_input()
                        i = input("\n\n\tanswer: ")
                        CTM.block_input()

                        if i != "exit!" and i != "exit" and i != "exit1":
                            CTM.clear_screen_backwards(down_to_row=8, delay=0)
                            NSP.play_accept_sound() if system_sound_level == 3 else None
                            url = "https://zenodo.org/record/5172857/files/wiki_morph.json?download=1"
                            normal = CA.download_database(url=url, directly_after_start=False)

                            if normal:
                                current_size = os.path.getsize("src/database/wiki_morph.json")
                                current_size = int(current_size / (1024 * 1024))
                                SDM.set_current_size(current_size)
                                database_version_description = SDM.get_database_version_description()
                                database_version_date = SDM.get_database_version_date()

                                CTM.clear_screen_backwards(down_to_row=8, delay=0)
                                print("\n\t\033[92mDownload completed sucessfully!\033[0m (" + str(current_size) +
                                      " MB)\n\tThe latest version of wikimorph is now installed on your device.")
                                NSP.play_request_sound() if system_sound_level >= 2 else None
                                print("\n\t\33[33mNote:\33[0m "
                                      "The version you just installed will be marked with its installation date"
                                      " as an identifier. \n\tBack in the settings menu you can add a short description "
                                      "\n\tfor additional information to the new installed version. "
                                      "\n\tThis is not mandatory and a description can also be added later or changed "
                                      "several times.")
                                time.sleep(3)
                                CTM.draw("\n\tPress \33[92menter\33[0m to load the database."
                                         "\n\tAfter that you will return to the settings menu automatically.",
                                         clear=False)
                                database_is_installed = True
                                CTM.unblock_input()
                                input()
                                CTM.block_input()
                                entries_list = CA.load_database(version=m2e_version)
                                CTM.clear_screen_backwards(down_to_row=5, delay=0)
                                print("\033[33m\n\t~ Settings Menu ~"
                                      "\n\t---------------------------------------------------"
                                      "----------------------------\033[0m")
                            else:
                                NSP.play_deny_sound() if system_sound_level == 3 else None
                        else:
                            CTM.draw("\n\t\33[33mReturning to overview...\33[0m")
                            NSP.play_deny_sound() if system_sound_level == 3 else None
                            time.sleep(2)

                    elif i == "2":

                        if CA.database_installation_confirmed():
                            description_set = False
                            while not description_set:
                                CTM.clear_screen_backwards(down_to_row=8, delay=0)
                                CA.display_settings_after_changes(setting=1, current_var="d1")
                                NSP.play_deny_sound() if system_sound_level >= 2 else None
                                CTM.unblock_input()
                                i = input("\n\tanswer: ")
                                CTM.block_input()
                                if len(i) > 25:
                                    CTM.draw("\n\t\33[91mInvalid entry:\33[0m description is too long!", clear=False)
                                    NSP.play_deny_sound() if system_sound_level >= 2 else None
                                    time.sleep(4)
                                else:
                                    CTM.draw("\n\t\33[92mEntry accepted!\33[0m", clear=False)
                                    NSP.play_accept_sound() if system_sound_level >= 2 else None
                                    time.sleep(2.5)
                                    database_version_description = i
                                    SDM.set_database_version_description(database_version_description)
                                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                                    CA.display_settings_after_changes(setting=1, current_var="d2",
                                                                      current_var_2=database_version_description)
                                    description_set = True
                                    time.sleep(4)
                        else:
                            SDM.set_database_version_date("")
                            SDM.set_database_version_description("")
                            CTM.unblock_input()
                            input()
                            CTM.block_input()
                            CTM.clear_screen_backwards(down_to_row=5, delay=0)
                            print("\033[33m\n\t~ Settings Menu ~"
                                  "\n\t---------------------------------------------------"
                                  "----------------------------\033[0m")

                    elif i == "3":

                        if CA.database_installation_confirmed():
                            CTM.clear_screen_backwards(down_to_row=8, delay=0)
                            CA.display_settings_after_changes(setting=1, current_var="r1")
                            NSP.play_deny_sound() if system_sound_level >= 2 else None

                            CTM.unblock_input()
                            i = input("\n\tanswer: ")
                            CTM.block_input()

                            if i != "exit!" and i != "exit" and i != "exit1":

                                NSP.play_accept_sound() if system_sound_level == 3 else None
                                SDM.set_database_version_date("")
                                SDM.set_database_version_description("")
                                if os.path.exists("src/database/wiki_morph.json"):
                                    os.remove("src/database/wiki_morph.json")

                                CTM.clear_screen_backwards(down_to_row=8, delay=0)
                                CA.display_settings_after_changes(setting=1, current_var="r2",
                                                                  current_var_2=database_version_date)
                                CTM.unblock_input()
                                input()
                                CTM.block_input()
                                CTM.draw("\n\t\33[92mReturning to overview...\33[0m")
                                time.sleep(2)

                            else:
                                CTM.draw("\n\t\33[33mReturning to settings menu...\33[0m")
                                NSP.play_deny_sound() if system_sound_level == 3 else None
                                time.sleep(2)
                        else:
                            SDM.set_database_version_date("")
                            SDM.set_database_version_description("")
                            database_version_description = SDM.get_database_version_description()
                            database_version_date = SDM.get_database_version_date()
                            CTM.unblock_input()
                            input()
                            CTM.block_input()
                            CTM.clear_screen_backwards(down_to_row=5, delay=0)
                            print("\033[33m\n\t~ Settings Menu ~"
                                  "\n\t--------------------------------------------------"
                                  "-----------------------------\033[0m")

                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 2

                if setting_ctrl == 2:
                    # setting 2 (term output diplomacy) ----------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(2, term_output_diplomacy)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mcontinue\33[0m without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\tType in \33[91mexit!\33[0m to \33[91mreturn to main menu\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()
                    if i == "1":
                        SDM.set_term_output_diplomacy("1")
                        term_output_diplomacy = SDM.get_term_output_diplomacy()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(2, term_output_diplomacy)
                        CTM.draw("\033[92m" + "\n\tOnly found terms will be considered!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_term_output_diplomacy("2")
                        term_output_diplomacy = SDM.get_term_output_diplomacy()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(2, term_output_diplomacy)
                        CTM.draw("\033[92m" + "\n\tOnly not found terms will be considered!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "3":
                        SDM.set_term_output_diplomacy("3")
                        term_output_diplomacy = SDM.get_term_output_diplomacy()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(2, term_output_diplomacy)
                        CTM.draw("\033[92m" + "\n\tAll terms will be considered!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 3

                if setting_ctrl == 3:
                    # setting 3 (Output line format) -------------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(3, oneline_output_format)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mcontinue\33[0m without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\tType in \33[91mexit!\33[0m to \33[91mreturn to main menu\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()
                    if i == "1":
                        SDM.set_one_line_output(True)
                        oneline_output_format = SDM.get_one_line_output()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(3, oneline_output_format)
                        CTM.draw("\033[92m" + "\n\tOutput will be printed in one-line format!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_one_line_output(False)
                        oneline_output_format = SDM.get_one_line_output()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(3, oneline_output_format)
                        CTM.draw("\033[92m" + "\n\tOutput will be printed in multi-line format!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 4

                if setting_ctrl == 4:
                    # setting 4 (headline-printing) --------------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(4, headline_printing)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mcontinue\33[0m without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\tType in \33[91mexit!\33[0m to \33[91mreturn to main menu\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()
                    if i == "1":
                        SDM.set_headline_printing("1")
                        headline_printing = SDM.get_headline_printing()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(4, headline_printing)
                        CTM.draw("\033[92m" + "\n\tHeadline will be printed only at top of excel!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_headline_printing("2")
                        headline_printing = SDM.get_headline_printing()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(4, headline_printing)
                        CTM.draw("\033[92m" + "\n\tHeadline will be printed for every new document in scan mode!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "3":
                        SDM.set_headline_printing("3")
                        headline_printing = SDM.get_headline_printing()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(4, headline_printing)
                        CTM.draw("\033[92m" + "\n\tHeadline will be printed for every new term!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 5

                if setting_ctrl == 5:
                    # setting 5 (alphabetical output) ------------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(5, alphabetical_output, abc_output_ascending)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mcontinue\33[0m without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\tType in \33[91mexit!\33[0m to \33[91mreturn to main menu\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()
                    if i == "1":
                        SDM.set_alphabetical_output(abc=True, asc=True)
                        alphabetical_output, abc_output_ascending = SDM.get_alphabetical_output()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(5, alphabetical_output, abc_output_ascending)
                        CTM.draw("\033[92m" + "\n\tOutput will be structured in ascending alphabetical order!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_alphabetical_output(abc=True, asc=False)
                        alphabetical_output, abc_output_ascending = SDM.get_alphabetical_output()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(5, alphabetical_output, abc_output_ascending)
                        CTM.draw("\033[92m" + "\n\tOutput will be structured in descending alphabetical order!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "3":
                        SDM.set_alphabetical_output(abc=False, asc=False)
                        alphabetical_output, abc_output_ascending = SDM.get_alphabetical_output()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(5, alphabetical_output, abc_output_ascending)
                        CTM.draw("\033[92m" + "\n\tOutput will not be structured at all!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 6

                if setting_ctrl == 6:
                    # setting 6 (auto scan filters) --------------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(6, auto_scan_filters)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mcontinue\33[0m without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\tType in \33[91mexit!\33[0m to \33[91mreturn to main menu\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()

                    if i == "1":
                        SDM.set_auto_scan_filters("Noun")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\t"Noun" is set as pos filter now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_auto_scan_filters("Verb")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\t"Verb" is set as pos filter now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "3":
                        SDM.set_auto_scan_filters("Adjective")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\t"Adjective" is set as pos filter now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "4":
                        SDM.set_auto_scan_filters("Adverb")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\t"Adverb" is set as pos filter now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "5":
                        SDM.set_auto_scan_filters("Preposition")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\t"Preposition" is set as pos filter now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "6":
                        SDM.set_auto_scan_filters("Phrase")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\t"Phrase" is set as pos filter now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "7":
                        SDM.set_auto_scan_filters("Noun, Verb, Adjective, Adverb, Preposition, Phrase")
                        auto_scan_filters = SDM.get_auto_scan_filters()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(6, auto_scan_filters)
                        CTM.draw("\033[92m" + '\n\tAll pos types will be considered now!\033[0m')
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 7

                if setting_ctrl == 7:
                    # setting 7 (output detail level) ------------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(7, output_detail_level)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mcontinue\33[0m without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\tType in \33[91mexit!\33[0m to \33[91mreturn to main menu\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()
                    if i == "1":
                        SDM.set_output_detail_level("1")
                        output_detail_level = SDM.get_output_detail_level()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(7, output_detail_level)
                        CTM.draw("\033[92m" + "\n\tOutput will cover term data only!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_output_detail_level("2")
                        output_detail_level = SDM.get_output_detail_level()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(7, output_detail_level)
                        CTM.draw("\033[92m" + "\n\tOutput will cover term data and morphology data!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "3":
                        SDM.set_output_detail_level("3")
                        output_detail_level = SDM.get_output_detail_level()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(7, output_detail_level)
                        CTM.draw("\033[92m" + "\n\tOutput will cover all data information!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    elif i == "exit" or i == "exit1" or i == "exit!":
                        setting_ctrl = 9
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 8

                if setting_ctrl == 8:
                    # setting 8 (system sound level) -------------------------------------------------------------------
                    CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    CA.display_settings(8, system_sound_level)
                    CTM.unblock_input()
                    i = input("\n\tType in the \33[33moption number\33[0m of the option you want to \33[33mchoose\33[0m"
                              " or\n\tpress \33[92menter\33[0m to \33[92mreturn to main menu\33[0m "
                              "without making changes."
                              "\n\tType in \33[94mb!\33[0m to navigate \33[94mbackwards\33[0m to the "
                              "\33[94mprevious setting\33[0m."
                              "\n\n\tOption number: ")
                    CTM.block_input()
                    if i == "1":
                        SDM.set_system_sound_level("1")
                        system_sound_level = SDM.get_system_sound_level()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(8, system_sound_level)
                        CTM.draw("\033[92m" + "\n\tNo sounds will be played!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "2":
                        SDM.set_system_sound_level("2")
                        system_sound_level = SDM.get_system_sound_level()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(8, system_sound_level)
                        CTM.draw("\033[92m" + "\n\tOnly notifications sounds will be played!\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "3":
                        SDM.set_system_sound_level("3")
                        system_sound_level = SDM.get_system_sound_level()
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                        CA.display_settings_after_changes(8, system_sound_level)
                        CTM.draw("\033[92m" + "\n\tNotification sounds and user feedback audio will be played!"
                                           "\033[0m")
                        NSP.play_deny_sound() if system_sound_level == 3 else None
                        time.sleep(4)
                    elif i == "b!" or i == "B!" or i == "b1" or i == "B1":
                        setting_ctrl -= 1
                        CTM.clear_screen_backwards(down_to_row=8, delay=0)
                    else:
                        setting_ctrl = 9

            CTM.clear_screen_backwards(down_to_row=8, delay=0)
            NSP.play_deny_sound() if system_sound_level >= 2 else None
            outro = "\033[92m" + "\n\n\tConfigurations were saved!" + "\033[0m" \
                    "\n\n\tReturning to main menu..."
            CTM.draw(outro, clear=False)
            time.sleep(4)
            print_main_menu_again = True

        else:
            # BASIC SEARCH ---------------------------------------------------------------------------------------------
            if CA.database_installation_confirmed():

                if CA.is_valid_input(i):

                    if not worksheet_generated:
                        workbook, worksheet, excel_row, hdlp_start = prepare_worksheet()
                        worksheet_generated = True

                    CTM.clear_screen_backwards(down_to_row=5)
                    CA.print_manual_search_headline()

                    open_excel_automatically = True
                    if ":" in i:
                        splitted_input = i.split(":")
                        term = splitted_input[0]
                        pos_filters = ""
                        for x in range(1, len(splitted_input)):
                            pos_filters += str(splitted_input[x]).capitalize()
                            if not x == len(splitted_input) - 1:
                                pos_filters += ", "
                        CTM.draw('\n\tSearching for term \33[33m' + term + '\33[0m '
                                 'with pos tag \33[33m(' + pos_filters + ')\33[0m...', clear=False)
                        time.sleep(2)
                    else:
                        pos_filters = "Noun, Verb, Adjective, Adverb, Preposition, Phrase"
                        term = i
                        CTM.draw('\n\tSearching for term \33[33m' + term + '\33[0m ...', clear=False)
                        time.sleep(2)

                    if term_output_diplomacy == 1:
                        worksheet, excel_row, log_output, \
                            hdlp_start, hdlp_doc = CA.search_and_output(worksheet=worksheet,
                                                                        excel_row=excel_row,
                                                                        pos_filters=pos_filters,
                                                                        term=term,
                                                                        entries_list=entries_list,
                                                                        only_found_terms=True,
                                                                        only_not_found_terms=False,
                                                                        multiline_output=not oneline_output_format,
                                                                        output_detail_level=output_detail_level,
                                                                        headline_printing=headline_printing,
                                                                        hdlp_start=hdlp_start,
                                                                        hdlp_doc=hdlp_doc)
                    elif term_output_diplomacy == 2:
                        worksheet, excel_row, log_output, \
                            hdlp_start, hdlp_doc = CA.search_and_output(worksheet=worksheet,
                                                                        excel_row=excel_row,
                                                                        pos_filters=pos_filters,
                                                                        term=term,
                                                                        entries_list=entries_list,
                                                                        only_found_terms=False,
                                                                        only_not_found_terms=True,
                                                                        multiline_output=not oneline_output_format,
                                                                        output_detail_level=output_detail_level,
                                                                        headline_printing=headline_printing,
                                                                        hdlp_start=hdlp_start,
                                                                        hdlp_doc=hdlp_doc)
                    else:
                        worksheet, excel_row, log_output, \
                            hdlp_start, hdlp_doc = CA.search_and_output(worksheet=worksheet,
                                                                        excel_row=excel_row,
                                                                        pos_filters=pos_filters,
                                                                        term=term,
                                                                        entries_list=entries_list,
                                                                        only_found_terms=False,
                                                                        only_not_found_terms=False,
                                                                        multiline_output=not oneline_output_format,
                                                                        output_detail_level=output_detail_level,
                                                                        headline_printing=headline_printing,
                                                                        hdlp_start=hdlp_start,
                                                                        hdlp_doc=hdlp_doc)

                    CTM.draw("\n\t\33[33mSaving results...\33[0m", clear=False)

                    log = open(log_title, "a", encoding="utf-8")
                    log.write("\n\n" + log_output)
                    log.close()

                    workbook.save(workbook_title)

                    print("\033[92m" + "\n\tDone!" + "\033[0m")
                    NSP.play_accept_sound() if system_sound_level >= 2 else None
                    time.sleep(1)
                else:
                    print_main_menu_again = True
            else:
                SDM.set_database_version_date("")
                SDM.set_database_version_description("")
                print_main_menu_again = True
                CTM.unblock_input()
                input()
                CTM.block_input()


# Program Start ----------------------------------------------------------------------------------------------
os.system("cls")
CTM.block_input()
set_console_fullscreen()
disable_resize_and_buttons()

time.sleep(1)

os.system("cls")
CA.print_opening(version=m2e_version, colour=True)
NSP.play_start_sound()
time.sleep(1)
CTM.draw("\n\tLoading components...")
time.sleep(1.5)
try:
    if SDM.get_first_start():
        set_system_variables_to_default()
        SDM.set_first_start(False)
    else:
        pass
except Exception:
    pass
# the problem with accessing first start will be solved in next exception handling


try:
    database_version_date = SDM.get_database_version_date()
    database_version_description = SDM.get_database_version_description()

    first_start = SDM.get_first_start()
    term_output_diplomacy = SDM.get_term_output_diplomacy()
    oneline_output_format = SDM.get_one_line_output()
    headline_printing = SDM.get_headline_printing()
    alphabetical_output, abc_output_ascending = SDM.get_alphabetical_output()
    auto_scan_filters = SDM.get_auto_scan_filters()
    output_detail_level = SDM.get_output_detail_level()
    system_sound_level = SDM.get_system_sound_level()

except Exception:
    print("\r\t\33[91mWarning: fatal system error detected!\33[0m")
    print("\n\tThis may be due to a problem with the savedata file.")
    time.sleep(3)
    print("\n\t\33[33mAll system variables will be set to default to solve the issue...\33[0m", end='', flush=True)
    set_system_variables_to_default()
    time.sleep(3)
    print("\r\t\33[33mRetrying accessing data...\33[0m")
    time.sleep(3)

    try:
        database_version_date = SDM.get_database_version_date()
        database_version_description = SDM.get_database_version_description()

        first_start = SDM.get_first_start()
        term_output_diplomacy = SDM.get_term_output_diplomacy()
        oneline_output_format = SDM.get_one_line_output()
        headline_printing = SDM.get_headline_printing()
        alphabetical_output, abc_output_ascending = SDM.get_alphabetical_output()
        auto_scan_filters = SDM.get_auto_scan_filters()
        output_detail_level = SDM.get_output_detail_level()
        system_sound_level = SDM.get_system_sound_level()

    except Exception:
        print("\r\t\33[91mWarning: fatal system error detected!\33[0m")
        print("\n\tThis may be due to a problem with the savedata file.")
        print("\n\t\33[91mUnfortunately the problem still remains. The program cannot be started so far.\33[0m"
              "\n\tPlease contact the developer for help:"
              "\n\n\t\33[92m" + supporter + "\n\t" + support_email + "\33[0m")
        time.sleep(3)
        print("\n\tYou can press \33[92menter\33[0m now to exit the program.")


check_database_installation()

formatted_date = CA.get_datetime()

log_name = CA.create_logfile(fd=formatted_date)

search_for_terms(log_name)
